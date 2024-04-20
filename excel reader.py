import openpyxl

def add_student():
    student_id = input("Enter student register number: ")
    name = input("Enter student name: ")
    age = input("Enter student age: ")
    grade = input("Enter student grade: ")
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Student ID", "Student Name", "Student Age", "Student Grade"])
    sheet.append([student_id, name, age, grade])
    workbook.save("students.xlsx")
    print("Student added successfully")

def login():
    name = input("Enter your name: ")
    student_id = input("Enter your ID: ")
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] == name and row[0] == student_id:
                print("Login successful!")
                return
        print("Login failed :(")
    except FileNotFoundError:
        print("No student records found. Please sign up.")

login_signup = input("1. Login\n2. Sign up\n")

if login_signup == "1":
    login()
elif login_signup == "2":
    add_student()
else:
    print("Invalid choice.")
