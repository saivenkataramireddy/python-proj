print("Welcome to SVIST Bank of India")

import openpyxl

def add_customer():
    customer_id = input("Enter customer ID: ")
    name = input("Enter customer name: ")
    age = input("Enter customer age: ")
    address = input("Enter customer address: ")
    balance = 0.0  # Initialize balance to 0.0

    try:
        workbook = openpyxl.load_workbook("customer.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == customer_id or row[1] == name:
                print("Customer already exists. Try another way.")
                return
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Customer ID", "Customer Name", "Customer Age", "Address", "Balance"])

    sheet.append([customer_id, name, age, address, balance])
    workbook.save("customer.xlsx")
    print("Account created successfully")

def login():
    name = input("Enter your name: ")
    customer_id = input("Enter your ID: ")

    try:
        workbook = openpyxl.load_workbook("customer.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=False):
            if row[1].value == name and row[0].value == customer_id:
                print("Login successful!")
                balance = row[4].value if len(row) > 4 else 0.0
                return balance, row, workbook
        print("Login failed :(")
    except FileNotFoundError:
        print("No customer records found. Please sign up.")
    return None, None, None

def show_balance(balance):
    print(f"Your balance is ${balance:.2f}")

def deposit():
    amount = float(input("Enter amount to deposit: "))
    if amount < 0:
        print("Invalid amount")
        return 0
    else:
        return amount

def withdraw(balance):
    amount = float(input("Enter amount to withdraw: "))
    if amount > balance:
        print("Insufficient funds")
        return 0
    elif amount <= 0:
        print("Amount must be greater than 0")
        return 0
    else:
        return amount

def update_balance(row, new_balance, workbook):
    row[4].value = new_balance
    workbook.save("customer.xlsx")

def delete_account(name, customer_id):
    try:
        workbook = openpyxl.load_workbook("customer.xlsx")
        sheet = workbook.active

        for i, row in enumerate(sheet.iter_rows(values_only=False)):
            if row[1].value == name and row[0].value == customer_id:
                sheet.delete_rows(i + 1)
                workbook.save("customer.xlsx")
                print("Account blocked (deleted) successfully!")
                return True

        print("Customer not found.")
        return False
    except FileNotFoundError:
        print("No customer records found.")
        return False

def main(balance, row, workbook):
    is_running = True
    while is_running:
        print("******************")
        print("\n1. Show Balance")
        print("******************")
        print("2. Deposit")
        print("******************")
        print("3. Withdraw")
        print("******************")
        print("4. Block Account")
        print("******************")
        print("5. Exit")
        print("******************")

        choice = input("Enter your choice: ")

        if choice == "1":
            show_balance(balance)
        elif choice == "2":
            balance += deposit()
            update_balance(row, balance, workbook)
        elif choice == "3":
            balance -= withdraw(balance)
            update_balance(row, balance, workbook)
        elif choice == "4":
            confirmation = input("Do you want to delete your account permanently? (Yes/No): ").lower()
            if confirmation == "yes":
                delete_account(row[1].value, row[0].value)
                break
            elif confirmation == "no":
                continue
            else:
                print("Invalid option. Please choose Yes or No.")
        elif choice == "5":
            is_running = False
        else:
            print("Invalid option")

    print("Thank you! Have a nice day!")

# Start of the program
login_signup = input("1. Create account:\n2. Login account:\n")

if login_signup == "1":
    add_customer()
elif login_signup == "2":
    balance, row, workbook = login()
    if balance is not None and row is not None:
        main(balance, row, workbook)
else:
    print("Invalid choice.")
